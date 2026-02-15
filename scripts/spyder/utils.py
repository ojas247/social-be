import sys
import os
import openpyxl
from datetime import datetime
from typing import Tuple, Optional, List
import csv
import calendar
import requests
from io import StringIO
from urllib.parse import urlparse
from google.cloud import storage
import google.generativeai as genai
from difflib import SequenceMatcher
from typing import List

sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
from app.services.datastore_services import fetch_entities_by_property
from app.utils.config import settings



genai.configure(api_key= settings.LLM_API_KEY)
# model = GenerativeModel("models/gemini-1.5-flash")
model = genai.GenerativeModel("models/gemini-2.0-flash")

def get_match_ratio(s1, s2):
    return SequenceMatcher(None, str(s1), str(s2)).ratio()


def get_last_date_of_month(date_str: str) -> str:
    """
    Converts 'Dec 2025' or 'December 2025' to '31-12-2025'.
    """
    try:
        # Parse the month and year (e.g., "Dec 2025")
        # %b is short month (Dec), %B is full month (December), %Y is 4-digit year
        try:
            date_obj = datetime.strptime(date_str.strip(), "%b %Y")
        except ValueError:
            date_obj = datetime.strptime(date_str.strip(), "%B %Y")
            
        month = date_obj.month
        year = date_obj.year
        
        # Get the last day of that specific month and year
        last_day = calendar.monthrange(year, month)[1]
        
        return f"{last_day:02d}-{month:02d}-{year}"
    except Exception:
        # Fallback to original string if parsing fails
        return date_str


def read_excel_cells(file_path: str, start_row: int, end_row: int) -> List[List]:
    """
    Read an Excel file and extract dates and values from columns B, C, and D starting from a given row.
    
    Args:
        file_path: Path to the Excel file
        start_row: Starting row number (default: 160)
        end_row: Ending row number (default: 160)
    Returns:
        List of lists (array of arrays) where:
        - First array: Dates (months) from column B
        - Second array: Values from column C (corresponding to dates)
        - Third array: Values from column D (corresponding to dates)
    """
    dates = []
    values_c = []
    values_d = []
    
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)
        
        # Get the active sheet (or specify sheet name if needed)
        sheet = workbook.active
        
        # Read from start_row downwards until we hit an empty cell in column B or end_row
        row = start_row
        while row <= end_row:
            # Read cell B (date)
            cell_b = sheet[f'B{row}']
            
            # Stop if cell B is empty
            if cell_b.value is None:
                break
            
            # Extract month from date
            month = None
            if isinstance(cell_b.value, datetime):
                # Extract month from datetime
                month = cell_b.value.strftime('%d-%m-%Y')  # e.g., "03-02-2024"
            else:
                # If it's already a string, use it as is
                month = str(cell_b.value)
            
            dates.append(month)
            
            # Read cell C
            cell_c = sheet[f'C{row}']
            value_c = cell_c.value if cell_c.value is not None else None
            values_c.append(value_c)
            
            # Read cell D
            cell_d = sheet[f'D{row}']
            value_d = cell_d.value if cell_d.value is not None else None
            values_d.append(value_d)
            
            row += 1
        
        # Close the workbook
        workbook.close()
        
        # Return as array of arrays: [dates, values_c, values_d]
        return [dates, values_c, values_d]
        
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        return [[], [], []]
    except Exception as e:
        print(f"Error reading Excel file: {str(e)}")
        return [[], [], []]


def fetch_csv_from_db(page_name):    
    # fetch_entities_by_property returns a LIST of dicts
    print("Started to fetch this page", page_name)
    entities = fetch_entities_by_property('Published_Data_v1', 'slugURL', page_name)
    # Check if the list is not empty before accessing
    if entities:
        print ("StartedV1")
        entity = entities[0] # Get the first matching record
        raw_report_url = entity.get('ReportUrl', '')
    else:
        print(f"Warning: No entity found for slug {page_name}")
    print('raw_report_url', raw_report_url)
    return raw_report_url


def convert_csv_to_arrayOfarray(csv_url: str) -> List[List[str]]:
    """
    Download a CSV file from a URL (bucket object) and convert it into an array of arrays (list of lists).
    
    Args:
        csv_url: URL of the CSV file stored in the bucket.
        
    Returns:
        A list of rows, where each row is a list of cell values (as strings).
    """
    table: List[List[str]] = []
    
    try:
        # Download the CSV file from the URL
        print(f"Downloading CSV from URL: {csv_url}")
        response = requests.get(csv_url, timeout=30)
        response.raise_for_status()  # Raise an exception for bad status codes
        
        # Decode the content (handle different encodings)
        content = response.content.decode('utf-8-sig')
        
        # Parse the CSV content using StringIO
        csv_file = StringIO(content)
        reader = csv.reader(csv_file)
        
        for row in reader:
            # Convert all cells to strings and append to table
            table.append([str(cell) for cell in row])
        
        print(f"Successfully converted CSV to array of arrays. Rows: {len(table)}")
        return table
        
    except requests.exceptions.RequestException as e:
        print(f"Error downloading CSV from URL: {str(e)}")
        return []
    except Exception as e:
        print(f"Error parsing CSV: {str(e)}")
        return []


def update_arrayOfarray(original_table: List[List], data_to_append: List[List]) -> List[List]:
    # Copy to avoid mutating original
    updated = [list(row) for row in original_table]
    
    # 1. Handle the Header (Dates)
    # data_to_append[0] example: ["Domestic Payment Frauds", "Dec 2025", "Jan 2026"]
    raw_dates = data_to_append[0][1:] 
    formatted_dates = [get_last_date_of_month(d) for d in raw_dates]
    updated[0].extend(formatted_dates)

    # 2. Process Data Rows
    # Skip the first list in data_to_append (it's the header)
    for new_row in data_to_append[1:]:
        new_label = new_row[0]  # e.g., "Volume (Lakh)"
        new_values = new_row[1:] # e.g., ["2.53999"]

        best_match_idx = -1
        highest_ratio = 0.0

        # Search for the best matching label in the original table
        # We skip updated[0] because it's the date header
        for i in range(1, len(updated)):
            existing_label = updated[i][0]
            ratio = get_match_ratio(new_label, existing_label)
            
            if ratio > highest_ratio:
                highest_ratio = ratio
                best_match_idx = i

        # 3. Append values if we found a decent match (threshold > 0.6)
        if best_match_idx != -1 and highest_ratio > 0.6:
            updated[best_match_idx].extend(new_values)
        else:
            print(f"Warning: No match found for label '{new_label}' (Best ratio: {highest_ratio})")
            # Optional: handle rows that exist in new data but not in original
            # updated.append(new_row) 

    return updated
  

def update_bucket_obj(updated_table: List[List], object_url: str) -> str:
    """
    Convert the updated table (array of arrays) to CSV and overwrite
    the corresponding object in Google Cloud Storage, whose public URL
    is given by raw_report_url.

    Args:
        updated_table: Array of arrays representing the CSV table.
        object_url: Public URL of the CSV object in GCS, e.g.:
            https://storage.googleapis.com/marketreports/Data/.../file.csv

    Returns:
        The gs:// path of the updated object.
    """
    if not updated_table or not object_url:
        raise ValueError("Both updated_table and object_url are required")

    # Convert table to CSV text
    csv_buffer = StringIO()
    writer = csv.writer(csv_buffer, lineterminator="\n")
    for row in updated_table:
        writer.writerow(row)
    csv_data = csv_buffer.getvalue()

    # Parse the URL to get bucket and blob path
    parsed = urlparse(object_url)

    # For URLs like: https://storage.googleapis.com/marketreports/Data/...
    if parsed.netloc == "storage.googleapis.com":
        path = parsed.path.lstrip("/")  # remove leading '/'
        parts = path.split("/", 1)
        if len(parts) != 2:
            raise ValueError(f"Could not parse bucket/blob from URL: {object_url}")
        bucket_name = parts[0]  # e.g., "marketreports"
        blob_name = f"Staging/{parts[1]}"  # e.g., "Staging/Data/file.csv"
    else:
        # Fallback: assume known bucket name and use whole path (without leading '/') as blob
        # e.g. assets.marketreports.in/Data/...  → bucket: marketreports, blob: Staging/Data/...
        bucket_name = "marketreports"
        path = parsed.path.lstrip("/")
        blob_name = f"Staging/{path}" if path else "Staging/"

    # Upload to GCS
    client = storage.Client()
    bucket = client.bucket(bucket_name)
    blob = bucket.blob(blob_name)
    blob.upload_from_string(csv_data, content_type="text/csv")

    return f"gs://{bucket_name}/{blob_name}"


def read_excel_row_ranges(file_path: str, row_ranges: List[Tuple[int, int]], sheet_name: str = None, 
                          column_ranges: List[Tuple[int, int]] = None) -> List[List[str]]:
    """
    Read specific row and column ranges from an Excel file and return them as a list of lists.
    
    Args:
        file_path: Path to the Excel file.
        row_ranges: List of tuples (min_row, max_row) where both are 0-indexed.
                    Example: [(2, 20), (90, 110)] reads rows 2-20 and 90-110.
        sheet_name: (Optional) Worksheet name, defaults to active sheet.
        column_ranges: List of tuples (min_col, max_col) where both are 0-indexed.
                       Example: [(0, 5), (10, 15)] reads columns 0-5 and 10-15.
                       If None, all columns are included.
    
    Returns:
        List of lists, where each inner list represents a row (as strings), filtered by column ranges.
    """
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        data_rows = []
        
        # Sort row ranges by min_row to process in order
        sorted_row_ranges = sorted(row_ranges, key=lambda x: x[0])
        
        # Process column ranges if provided
        column_indices = None
        if column_ranges:
            column_indices = set()
            sorted_col_ranges = sorted(column_ranges, key=lambda x: x[0])
            for min_col, max_col in sorted_col_ranges:
                column_indices.update(range(min_col, max_col + 1))
        
        # Read all rows first (since we need to access by index)
        all_rows = list(sheet.iter_rows(values_only=True))
        
        # Extract rows from each range
        for min_row, max_row in sorted_row_ranges:
            for i in range(min_row, min(max_row + 1, len(all_rows))):
                row = all_rows[i]
                str_row = [str(cell) if cell is not None else "" for cell in row]
                
                # Filter columns if column_ranges is provided
                if column_indices is not None:
                    filtered_row = [str_row[j] for j in range(len(str_row)) if j in column_indices]
                    data_rows.append(filtered_row)
                else:
                    data_rows.append(str_row)
        
        wb.close()
        return data_rows
    except Exception as ex:
        print(f"Error reading Excel row/column ranges: {ex}")
        return []


def format_rows_to_context(data_rows: List[List[str]]) -> str:
    """
    Convert a list of row lists into a tab-separated text format for LLM context.
    
    Args:
        data_rows: List of lists, where each inner list represents a row.
    
    Returns:
        A string with rows separated by newlines and cells separated by tabs.
    """
    context_lines = []
    for row in data_rows:
        context_lines.append("\t".join(row))
    return "\n".join(context_lines)


def excel_to_gemini_context(file_path: str, custom_prompt: str, sheet_name: str = None, 
                            row_ranges: List[Tuple[int, int]] = None, 
                            column_ranges: List[Tuple[int, int]] = None,
                            min_rows: int = None, max_rows: int = None) -> str:
    """
    Read an .xlsx file and prepare context for a LLM (like Gemini) to extract a specific value,
    using a custom prompt. The function can load specific row and column ranges or a single range.

    Args:
        file_path: Path to the Excel file.
        custom_prompt: Custom text prompt, describing what data/value to fetch.
        sheet_name: (Optional) Worksheet name, defaults to active sheet.
        row_ranges: List of tuples (min_row, max_row) for specific ranges. 
                    Example: [(2, 20), (90, 110)] reads rows 2-20 and 90-110.
                    Takes precedence over min_rows/max_rows if provided.
        column_ranges: List of tuples (min_col, max_col) for specific column ranges.
                       Example: [(0, 5), (10, 15)] reads columns 0-5 and 10-15.
                       If None, all columns are included.
        min_rows: Minimum row index (0-indexed) for single range (backward compatibility).
        max_rows: Maximum row index (0-indexed) for single range (backward compatibility).
        Both rows and columns are 0-indexed

    Returns:
        A string containing tabular data context and the custom prompt, suitable for a chat-based LLM.
    """
    try:
        # Determine which ranges to use
        if row_ranges:
            # Use provided row ranges
            data_rows = read_excel_row_ranges(file_path, row_ranges, sheet_name, column_ranges)
        elif min_rows is not None and max_rows is not None:
            # Use single range for backward compatibility
            data_rows = read_excel_row_ranges(file_path, [(min_rows, max_rows)], sheet_name, column_ranges)
        else:
            # Default: first 30 rows
            data_rows = read_excel_row_ranges(file_path, [(0, 30)], sheet_name, column_ranges)
        
        # Convert rows to text format
        table_context = format_rows_to_context(data_rows)
        
        gemini_instruction = (
            f"Below is a preview of the Excel sheet in tabular form (rows are tab-separated):\n"
            f"{table_context}\n\n"
            f"Prompt: {custom_prompt.strip()}\n"
            f"Please analyze the table and provide only the exact requested data in this format list-of-list format like below: [[Name of Table, Date1, Date2, Date3....],[Item1, Value1, Value2, Value3....], [Item2, ValueX, ValueY, ValueZ....]...so on]"
        )
        return gemini_instruction
    except Exception as ex:
        return f"Error reading Excel for Gemini context: {ex}"


def send_prompt_to_gemini(prompt: str) -> str:
    """
    Sends a prompt to Gemini LLM API and returns the response as a string.
    """
    response = model.generate_content(prompt)
    return response.text
