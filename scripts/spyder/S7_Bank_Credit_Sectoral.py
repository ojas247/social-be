from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
import os
import re
import sys

import openpyxl
import pandas as pd
import requests
from playwright.sync_api import sync_playwright
from urllib.parse import urljoin

from utils import fetch_item_names_from_TimeSeriesData

sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
from app.services.datastore_services import update_Datastore


RBI_LIST_URL = "https://rbi.org.in/Scripts/Data_Sectoral_Deployment.aspx"
STATEMENTS_LINK_PATTERN = re.compile(r"Statements I and II\.?", re.I)
RELEASE_LINK_PATTERN = re.compile(r"Sectoral Deployment of Bank Credit", re.I)
SINGLE_DATE_PATTERN = re.compile(r"\d{1,2}\.[A-Za-z]{3,9},\s*\d{4}")
VALUE_COLUMN_CANDIDATES = ("F", "E", "D")
DATE_HEADER_ROW = 4
FUZZY_THRESHOLD = 0.72


MASTER_LIST = [
    "Agriculture and Allied Activities",
    "Micro and Small Industry",
    "Medium Industry",
    "Large Industry",
    "Transport Operators Services",
    "Computer Software Services",
    "Tourism, Hotels and Restaurants Services",
    "Shipping Services",
    "Aviation Services",
    "Professional Services Services",
    "Wholesale Trade1 Services",
    "Retail Trade Services",
    "Commercial Real Estate Services",
    "Non-Banking Financial Companies",
    "Housing Finance Companies (NBFC)",
    "Public Financial Institutions (NBFC)",
    "Other Services",
    "Personal Loans -  Consumer Durables",
    "Personal Loans - Housing (Including Priority Sector Housing)",
    "Personal Loans -  Advances against Fixed Deposits",
    "Personal Loans -  Advances to Individuals against share, bonds, etc",
    "Personal Loans -  Credit Card Outstanding",
    "Personal Loans -  Education",
    "Personal Loans -  Vehicle Loans",
    "Personal Loans -  Loans against gold jewellery",
    "Other Personal Loans",
]

def get_match_ratio(s1, s2) -> float:
    return SequenceMatcher(None, str(s1), str(s2)).ratio()


def _to_whole_number(value) -> int:
    """Round numeric Excel values to 0 decimal places for StagingData."""
    return int(round(float(value)))



def _month_number(month: str) -> int:
    """Accept month as name ('May'/'Feb') or numeric ('5'/'05')."""
    raw = month.strip()
    if raw.isdigit():
        value = int(raw)
        if 1 <= value <= 12:
            return value
        raise ValueError(f"Invalid month number: {month}")

    for fmt in ("%B", "%b"):
        try:
            return datetime.strptime(raw.title(), fmt).month
        except ValueError:
            continue
    raise ValueError(f"Unrecognized month: {month}")


def _dates_in_cell(cell_value) -> list[str]:
    """Return date strings like '31.May,2026' found in a header cell."""
    if cell_value is None:
        return []
    if isinstance(cell_value, datetime):
        return [cell_value.strftime("%d.%b,%Y")]
    return SINGLE_DATE_PATTERN.findall(str(cell_value).strip())


def _is_single_date_header(cell_value) -> bool:
    """True when the cell contains exactly one d.Mon,YYYY date."""
    return len(_dates_in_cell(cell_value)) == 1


def _month_from_date_header(cell_value) -> int:
    """Parse month from a single-date header like '31.May,2026'."""
    dates = _dates_in_cell(cell_value)
    if len(dates) != 1:
        raise ValueError(f"Expected one date in header cell, got {cell_value!r}")

    text = dates[0]
    for fmt in ("%d.%b,%Y", "%d.%B,%Y"):
        try:
            return datetime.strptime(text, fmt).month
        except ValueError:
            continue
    raise ValueError(f"Could not parse month from date header: {cell_value!r}")


def _select_value_column(ws) -> tuple[str, int, object]:
    """
    Pick the value column from F, then E, then D.
    Header row must contain exactly one date in d.Mon,YYYY format.
    """
    for col_letter in VALUE_COLUMN_CANDIDATES:
        header_value = ws[f"{col_letter}{DATE_HEADER_ROW}"].value
        if _is_single_date_header(header_value):
            col_idx = ord(col_letter) - ord("A")
            return col_letter, col_idx, header_value

    raise ValueError(
        "No value column found in F/E/D with a single date header "
        f"on row {DATE_HEADER_ROW}"
    )


def _normalize_text(text: str) -> str:
    s = str(text).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _normalize_excel_label(label: str) -> str:
    """Strip row numbering, footnotes, and noisy suffixes from sheet labels."""
    s = _normalize_text(label)
    s = re.sub(r"^\d+(\.\d+)*\.?\s*", "", s)
    s = re.sub(r"([a-z)])\d+\b", r"\1", s)
    s = re.sub(r",?\s*of which,?.*$", "", s)
    s = re.sub(r"[()]", " ", s)
    s = re.sub(r"\s+", " ", s).strip(" ,.-")
    return s


def _master_variants(master: str) -> list[str]:
    """Build comparable forms of a master label for fuzzy matching."""
    base = _normalize_text(master)
    base = re.sub(r"^personal loans\s*-\s*", "", base)
    variants = {base}

    stripped = re.sub(r"\s+(industry|services)$", "", base)
    variants.add(stripped)
    variants.add(re.sub(r"\s*\(nbfcs?\)$", "", base))
    variants.add(re.sub(r"\s*\(nbfcs?\)$", "", stripped))

    softened = {re.sub(r"[()]", " ", v) for v in variants}
    softened = {re.sub(r"\s+", " ", v).strip(" ,.-") for v in softened}
    return [v for v in softened if v]


def _match_score(master: str, excel_label: str) -> float:
    excel_norm = _normalize_excel_label(excel_label)
    if not excel_norm:
        return 0.0

    best = 0.0
    for variant in _master_variants(master):
        best = max(best, get_match_ratio(variant, excel_norm))
        if variant == excel_norm:
            best = max(best, 1.0)
        elif len(variant) >= 8 and (
            excel_norm.startswith(variant) or variant.startswith(excel_norm)
        ):
            best = max(best, 0.95)
    return best


def match_master_to_excel(
    excel_rows: list[dict],
    master_item_names: list[str],
) -> pd.DataFrame:
    """Keep only rows that fuzzy-match master_item_names."""
    candidates = [
        row for row in excel_rows
        if row.get("Sector") and row.get("Value") is not None
        and not isinstance(row["Value"], str)
    ]

    used_indices: set[int] = set()
    matched = []

    for master in master_item_names:
        best_idx = None
        best_score = 0.0
        for i, row in enumerate(candidates):
            if i in used_indices:
                continue
            score = _match_score(master, row["Sector"])
            if score > best_score:
                best_score = score
                best_idx = i

        if best_idx is None or best_score < FUZZY_THRESHOLD:
            continue

        used_indices.add(best_idx)
        matched.append({
            "Sector": master,
            "Value": _to_whole_number(candidates[best_idx]["Value"]),
        })

    return pd.DataFrame(matched)


def download_sectoral_excel(row: int = 1, download_dir: Path | None = None) -> Path:
    """
    Scrape RBI sectoral deployment Excel from the listing page.

    row=1 selects the latest release link (top row); row=2 is the second, etc.
    """
    if row < 1:
        raise ValueError("row must be >= 1 (1 = latest release)")

    save_dir = download_dir or (Path(__file__).resolve().parent / "downloads")
    save_dir.mkdir(parents=True, exist_ok=True)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(RBI_LIST_URL, wait_until="domcontentloaded", timeout=120_000)

        release_links = page.locator("table").first.locator("a").filter(
            has_text=RELEASE_LINK_PATTERN
        )
        count = release_links.count()
        if count == 0:
            browser.close()
            raise RuntimeError("No sectoral deployment release links found on RBI page")
        if row > count:
            browser.close()
            raise IndexError(
                f"Requested row {row} is out of bounds. Page has {count} release row(s)."
            )

        selected_text = release_links.nth(row - 1).inner_text().strip()
        with page.expect_navigation(wait_until="domcontentloaded", timeout=120_000):
            release_links.nth(row - 1).click()

        detail_url = page.url
        statements_links = page.locator("a").filter(
            has_text=STATEMENTS_LINK_PATTERN
        )
        if statements_links.count() == 0:
            browser.close()
            raise RuntimeError(
                f'"Statements I and II." link not found on {detail_url}'
            )

        download_href = statements_links.first.get_attribute("href")
        browser.close()

    if not download_href:
        raise RuntimeError(
            f'"Statements I and II." link on {detail_url} has no href'
        )

    download_url = urljoin(detail_url, download_href)
    response = requests.get(download_url, timeout=120)
    response.raise_for_status()

    filename = Path(download_url.split("?")[0]).name or "sectoral_bank_credit.xlsx"
    xlsx_path = save_dir / filename
    xlsx_path.write_bytes(response.content)
    print(f"Downloaded {selected_text!r} -> {xlsx_path}")
    return xlsx_path


def main():
    ts_kind = "TimeSeriesData"
    month = "May"
    date = f"{month} 2026"
    scriptID = "S7"
    staging_kind = "StagingData_v1"
    granularity = "Monthly"
    property_name = "dataName"
    dataName = "Sectoral Deployment of Bank Credit"
    release_row = 1  # 1 = latest release on RBI page

    master_item_names = fetch_item_names_from_TimeSeriesData(
        ts_kind, property_name, dataName
    )
    xlsx_path = download_sectoral_excel(row=release_row)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    value_col, value_col_idx, header_value = _select_value_column(ws)
    print(f"Using column {value_col} (header: {header_value!r})")

    header_month = _month_from_date_header(header_value)
    input_month = _month_number(month)
    if header_month != input_month:
        raise ValueError(
            f"Month mismatch: input is {month!r} but "
            f"{value_col}{DATE_HEADER_ROW} has {header_value!r}"
        )

    rows = []
    for row in ws.iter_rows(min_col=1, max_col=6, values_only=True):
        col_a = row[0]
        col_val = row[value_col_idx] if len(row) > value_col_idx else None
        if col_a is None and col_val is None:
            continue
        rows.append({"Sector": col_a, "Value": col_val})

    df = match_master_to_excel(rows, master_item_names)
    print(df.to_string(index=False))
    missing = [m for m in master_item_names if m not in set(df["Sector"])]
    if missing:
        print(f"\nUnmatched master entries ({len(missing)}):")
        for m in missing:
            print(f"  - {m}")

    parsed = {row["Sector"]: row["Value"] for row in df.to_dict(orient="records")}
    update_Datastore(
        parsed,
        date,
        granularity,
        scriptID,
        str(xlsx_path),
        dataName,
        staging_kind,
    )


if __name__ == "__main__":
    main()
