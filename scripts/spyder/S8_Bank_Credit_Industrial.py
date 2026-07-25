from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
import os
import re
import sys

import openpyxl
import pandas as pd
import requests
from playwright.sync_api import sync_playwright
from urllib.parse import urljoin

from utils import fetch_item_names_from_TimeSeriesData

sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
from app.services.datastore_services import update_Datastore


RBI_LIST_URL = "https://rbi.org.in/Scripts/Data_Sectoral_Deployment.aspx"
STATEMENTS_LINK_PATTERN = re.compile(r"Statements I and II\.?", re.I)
RELEASE_LINK_PATTERN = re.compile(r"Sectoral Deployment of Bank Credit", re.I)
SINGLE_DATE_PATTERN = re.compile(r"\d{1,2}\.[A-Za-z]{3,9},\s*\d{4}")
VALUE_COLUMN_CANDIDATES = ("F", "E", "D")
DATE_HEADER_ROW = 4
FUZZY_THRESHOLD = 0.72
STATEMENT_SHEET_NAME = "Statement 2"


# Excel label (Statement 2 col A) -> Datastore entity `item` name
MASTER_EXCEL_MAPPER = {
  "2.1. Mining and Quarrying (incl. Coal)": "Mining and Quarrying (incl. Coal)",
  "2.2.1. Sugar": "Food Processing - Sugar",
  "2.2.2. Edible Oils and Vanaspati": "Food Processing - Edible Oils and Vanaspati",
  "2.2.3. Tea": "Food Processing - Tea",
  "2.2.4. Others": "Food Processing - Others",
  "2.3. Beverage and Tobacco": "Beverage and Tobacco",
  "2.4.1. Cotton Textiles": "Cotton Textiles",
  "2.4.2. Jute Textiles": "Jute Textiles",
  "2.4.3. Man-Made Textiles": "Man-Made Textiles",
  "2.4.4. Other Textiles": "Other Textiles",
  "2.5. Leather and Leather Products": "Leather and Leather Products",
  "2.6. Wood and Wood Products": "Wood and Wood Products",
  "2.7. Paper and Paper Products": "Paper and Paper Products",
  "2.8. Petroleum, Coal Products and Nuclear Fuels": "Petroleum, Coal Products and Nuclear Fuels",
  "2.9.1. Fertiliser": "Chemicals - Fertiliser",
  "2.9.2. Drugs and Pharmaceuticals": "Chemicals - Drugs & Pharmaceuticals",
  "2.9.3. Petro Chemicals": "Chemicals - Petro Chemicals",
  "2.9.4. Others": "Chemicals - Others",
  "2.10. Rubber, Plastic and their Products": "Rubber, Plastic and their Products",
  "2.11. Glass and Glassware": "Glass and Glassware",
  "2.12. Cement and Cement Products": "Cement and Cement Products",
  "2.13.1. Iron and Steel": "Iron and Steel",
  "2.13.2. Other Metal and Metal Product": "Other Metal and Metal Product",
  "2.14.1. Electronics": "Engineering - Electronics",
  "2.14.2. Others": "Engineering - Others",
  "2.15. Vehicles, Vehicle Parts and Transport Equipment": "Vehicles, Vehicle Parts and Transport Equipment",
  "2.16. Gems and Jewellery": "Gems and Jewellery",
  "2.17. Construction": "Construction",
  "2.18.1. Power": "Infrastructure - Power",
  "2.18.2. Telecommunications": "Infrastructure - Telecommunications",
  "2.18.3. Roads": "Infrastructure - Roads",
  "2.18.4. Airports": "Infrastructure - Airports",
  "2.18.5. Ports": "Infrastructure - Ports",
  "2.18.6. Railways (other than Indian Railways)": "Infrastructure - Railways (other than Indian Railways)",
  "2.18.7. Other Infrastructure": "Other Infrastructure",
  "2.19. Other Industries": "Other Industries"
}

def get_match_ratio(s1, s2) -> float:
    return SequenceMatcher(None, str(s1), str(s2)).ratio()


def _to_whole_number(value) -> int:
    """Round numeric Excel values to 0 decimal places for StagingData."""
    return int(round(float(value)))



def _month_number(month: str) -> int:
    """Accept month as name ('May'/'Feb') or numeric ('5'/'05')."""
    raw = month.strip()
    if raw.isdigit():
        value = int(raw)
        if 1 <= value <= 12:
            return value
        raise ValueError(f"Invalid month number: {month}")

    for fmt in ("%B", "%b"):
        try:
            return datetime.strptime(raw.title(), fmt).month
        except ValueError:
            continue
    raise ValueError(f"Unrecognized month: {month}")


def _dates_in_cell(cell_value) -> list[str]:
    """Return date strings like '31.May,2026' found in a header cell."""
    if cell_value is None:
        return []
    if isinstance(cell_value, datetime):
        return [cell_value.strftime("%d.%b,%Y")]
    return SINGLE_DATE_PATTERN.findall(str(cell_value).strip())


def _is_single_date_header(cell_value) -> bool:
    """True when the cell contains exactly one d.Mon,YYYY date."""
    return len(_dates_in_cell(cell_value)) == 1


def _month_from_date_header(cell_value) -> int:
    """Parse month from a single-date header like '31.May,2026'."""
    dates = _dates_in_cell(cell_value)
    if len(dates) != 1:
        raise ValueError(f"Expected one date in header cell, got {cell_value!r}")

    text = dates[0]
    for fmt in ("%d.%b,%Y", "%d.%B,%Y"):
        try:
            return datetime.strptime(text, fmt).month
        except ValueError:
            continue
    raise ValueError(f"Could not parse month from date header: {cell_value!r}")


def _select_value_column(ws) -> tuple[str, int, object]:
    """
    Pick the value column from F, then E, then D.
    Header row must contain exactly one date in d.Mon,YYYY format.
    """
    for col_letter in VALUE_COLUMN_CANDIDATES:
        header_value = ws[f"{col_letter}{DATE_HEADER_ROW}"].value
        if _is_single_date_header(header_value):
            col_idx = ord(col_letter) - ord("A")
            return col_letter, col_idx, header_value

    raise ValueError(
        "No value column found in F/E/D with a single date header "
        f"on row {DATE_HEADER_ROW}"
    )


def _get_statement_sheet(wb, sheet_name: str = STATEMENT_SHEET_NAME):
    """
    Return the named worksheet (e.g. 'Statement 2'), even when it is not active.
    Falls back to a case-insensitive / whitespace-tolerant name match.
    """
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    target = sheet_name.strip().lower()
    for name in wb.sheetnames:
        if name.strip().lower() == target:
            return wb[name]

    # e.g. "Statement 2" vs "Statement2" / "Statements 2"
    compact_target = re.sub(r"\s+", "", target)
    for name in wb.sheetnames:
        if re.sub(r"\s+", "", name.strip().lower()) == compact_target:
            return wb[name]

    raise KeyError(
        f"Worksheet {sheet_name!r} not found. Available sheets: {wb.sheetnames}"
    )


def _normalize_text(text: str) -> str:
    s = str(text).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _normalize_excel_label(label: str, *, keep_section_numbers: bool = False) -> str:
    """Normalize sheet / mapper labels for fuzzy matching."""
    s = _normalize_text(label)
    if not keep_section_numbers:
        # Leading section numbers: "3.9.1. Housing..." / "1. Agriculture..."
        s = re.sub(r"^\d+(\.\d+)*\.?\s*", "", s)
    # Trailing footnote digit glued to a word: jewellery4, Services3
    s = re.sub(r"([a-z)])\d+\b", r"\1", s)
    s = re.sub(r",?\s*of which,?.*$", "", s)
    s = re.sub(r"[()]", " ", s)
    s = re.sub(r"\s+", " ", s).strip(" ,.-")
    return s


def _match_score(mapper_key: str, excel_label: str) -> float:
    """
    Score how well an Excel row label matches a MASTER_EXCEL_MAPPER key.
    Keeps section numbers (2.9.4 vs 2.2.4) so 'Others' rows stay distinct.
    """
    key_norm = _normalize_excel_label(mapper_key, keep_section_numbers=True)
    excel_norm = _normalize_excel_label(excel_label, keep_section_numbers=True)
    if not key_norm or not excel_norm:
        return 0.0

    best = get_match_ratio(key_norm, excel_norm)
    if key_norm == excel_norm:
        return 1.0
    if len(key_norm) >= 8 and (
        excel_norm.startswith(key_norm) or key_norm.startswith(excel_norm)
    ):
        best = max(best, 0.95)
    return best


def match_excel_via_mapper(
    excel_rows: list[dict],
    mapper: dict[str, str] = MASTER_EXCEL_MAPPER,
) -> tuple[pd.DataFrame, list[dict]]:
    """
    Map Excel rows to datastore item names using MASTER_EXCEL_MAPPER.

    Mapper keys = labels as they appear in Excel.
    Mapper values = entity `item` / property names for StagingData.
    """
    candidates = [
        row for row in excel_rows
        if row.get("Sector") and row.get("Value") is not None
        and not isinstance(row["Value"], str)
    ]

    used_indices: set[int] = set()
    matched = []
    unmatched = []

    for excel_key, datastore_item in mapper.items():
        best_idx = None
        best_score = 0.0
        for i, row in enumerate(candidates):
            if i in used_indices:
                continue
            score = _match_score(excel_key, row["Sector"])
            if score > best_score:
                best_score = score
                best_idx = i

        if best_idx is None or best_score < FUZZY_THRESHOLD:
            # Closest label across all rows (including already-used) for diagnostics
            closest_label = None
            closest_score = 0.0
            for row in candidates:
                score = _match_score(excel_key, row["Sector"])
                if score > closest_score:
                    closest_score = score
                    closest_label = row["Sector"]
            unmatched.append({
                "ExcelKey": excel_key,
                "DatastoreItem": datastore_item,
                "MatchScore": round(closest_score, 3),
                "ClosestExcel": closest_label,
            })
            continue

        used_indices.add(best_idx)
        matched.append({
            "Sector": datastore_item,
            "Value": _to_whole_number(candidates[best_idx]["Value"]),
            "ExcelSector": candidates[best_idx]["Sector"],
            "MatchScore": round(best_score, 3),
        })

    return pd.DataFrame(matched), unmatched


def download_sectoral_excel(row: int = 1, download_dir: Path | None = None) -> Path:
    """
    Scrape RBI sectoral deployment Excel from the listing page.

    row=1 selects the latest release link (top row); row=2 is the second, etc.
    """
    if row < 1:
        raise ValueError("row must be >= 1 (1 = latest release)")

    save_dir = download_dir or (Path(__file__).resolve().parent / "downloads")
    save_dir.mkdir(parents=True, exist_ok=True)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(RBI_LIST_URL, wait_until="domcontentloaded", timeout=120_000)

        release_links = page.locator("table").first.locator("a").filter(
            has_text=RELEASE_LINK_PATTERN
        )
        count = release_links.count()
        if count == 0:
            browser.close()
            raise RuntimeError("No sectoral deployment release links found on RBI page")
        if row > count:
            browser.close()
            raise IndexError(
                f"Requested row {row} is out of bounds. Page has {count} release row(s)."
            )

        selected_text = release_links.nth(row - 1).inner_text().strip()
        with page.expect_navigation(wait_until="domcontentloaded", timeout=120_000):
            release_links.nth(row - 1).click()

        detail_url = page.url
        statements_links = page.locator("a").filter(
            has_text=STATEMENTS_LINK_PATTERN
        )
        if statements_links.count() == 0:
            browser.close()
            raise RuntimeError(
                f'"Statements I and II." link not found on {detail_url}'
            )

        download_href = statements_links.first.get_attribute("href")
        browser.close()

    if not download_href:
        raise RuntimeError(
            f'"Statements I and II." link on {detail_url} has no href'
        )

    download_url = urljoin(detail_url, download_href)
    response = requests.get(download_url, timeout=120)
    response.raise_for_status()

    filename = Path(download_url.split("?")[0]).name or "sectoral_bank_credit.xlsx"
    xlsx_path = save_dir / filename
    xlsx_path.write_bytes(response.content)
    print(f"Downloaded {selected_text!r} -> {xlsx_path}")
    return xlsx_path


def main():
    ts_kind = "TimeSeriesData"
    month = "Jan"
    date = f"{month} 2026"
    scriptID = "S8"
    staging_kind = "StagingData_v1"
    granularity = "Monthly"
    property_name = "dataName"
    dataName = "Industry-wise Deployment of Bank Credit"
    release_row = 5 # 1 = latest release on RBI page

    master_item_names = fetch_item_names_from_TimeSeriesData(
        ts_kind, property_name, dataName
    )
    xlsx_path = download_sectoral_excel(row=release_row)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = _get_statement_sheet(wb, STATEMENT_SHEET_NAME)
    print(f"Using sheet {ws.title!r} (workbook sheets: {wb.sheetnames})")

    value_col, value_col_idx, header_value = _select_value_column(ws)
    print(f"Using column {value_col} (header: {header_value!r})")

    header_month = _month_from_date_header(header_value)
    input_month = _month_number(month)
    if header_month != input_month:
        raise ValueError(
            f"Month mismatch: input is {month!r} but "
            f"{value_col}{DATE_HEADER_ROW} has {header_value!r}"
        )

    rows = []
    for row in ws.iter_rows(min_col=1, max_col=6, values_only=True):
        col_a = row[0]
        col_val = row[value_col_idx] if len(row) > value_col_idx else None
        if col_a is None and col_val is None:
            continue
        rows.append({"Sector": col_a, "Value": col_val})

    df, unmatched = match_excel_via_mapper(rows, MASTER_EXCEL_MAPPER)
    # StagingData uses mapper values (datastore item names) as keys.
    if not df.empty:
        print(df[["Sector", "Value"]].to_string(index=False))
    else:
        print("(no matched rows)")

    if unmatched:
        print(
            f"\nUnmatched MASTER_EXCEL_MAPPER entries ({len(unmatched)}) "
            f"[threshold={FUZZY_THRESHOLD}]:"
        )
        for u in unmatched:
            print(
                f"  - excel_key={u['ExcelKey']!r} -> item={u['DatastoreItem']!r} "
                f"| score={u['MatchScore']:.3f} | closest={u['ClosestExcel']!r}"
            )

    mapped_items = set(df["Sector"]) if not df.empty else set()
    missing_in_db = [
        item for item in MASTER_EXCEL_MAPPER.values()
        if item in mapped_items and item not in set(master_item_names)
    ]
    if missing_in_db:
        print(
            f"\nMapped items not found in TimeSeriesData ({len(missing_in_db)}):"
        )
        for item in missing_in_db:
            print(f"  - {item}")

    parsed = {row["Sector"]: row["Value"] for row in df.to_dict(orient="records")}
    update_Datastore(
        parsed,
        date,
        granularity,
        scriptID,
        str(xlsx_path),
        dataName,
        staging_kind,
    )


if __name__ == "__main__":
    main()
